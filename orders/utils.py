from openpyxl import Workbook


def generate_report(data: list[dict]) -> Workbook:
    """
    Генерирует отчет в формате Excel с отдельными листами для каждой модели и версии
    с количеством произведенных единиц роботов за последнюю неделю.

    :param data: Список словарей, содержащих информацию о моделях.
    :type data: list[dict]

    :return: Объект книги Excel, содержащий сгенерированный отчет.
    :rtype: Workbook
    """
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    for model_data in data:
        model_name = model_data['model']
        model_version = model_data['version']
        week_amount = model_data['week_amount']

        sheet_exists = False
        for sheet in wb.sheetnames:
            if sheet == model_name:
                sheet_exists = True
                break

        if not sheet_exists:
            ws = wb.create_sheet(title=model_name)
            ws.append(('Модель', 'Версия', 'Количество за неделю'))
        else:
            ws = wb[model_name]

        ws.append((model_name, model_version, week_amount))
    return wb
